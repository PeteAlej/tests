# dev_test_v1

import fnmatch
import os, shutil
import time
import keyboard
# Para utilizar el módulo de watchdog es necesario el módulo de PyYAML y posiblemente LibYAML, argh, argparse y pathtools
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import openpyxl as xl

#dir = os.path.dirname(__file__)
#xls_path = os.path.join(dir,'Processed\\')

# Cambiar los paths para hacer pruebas en otras computadoras
src_path = r"C:/Users/Peter/Documents/GitHub/tests/dev_test/"					# Directorio donde se van a "crear" los archivos
xls_path = r"C:/Users/Peter/Documents/GitHub/tests/dev_test/Processed/"			# Directorio donde se van a mover los archivos .xls*
nonxls_path = r"C:/Users/Peter/Documents/GitHub/tests/dev_test/Not_applicable/"	# Directorio donde se van a mover los archivos != .xls*
directory = "."

class watcher:			# Inicialización de la clase "watcher", la cual establece las instancias con los argumentos ingresados.
	def __init__(self, directory = directory, handler = FileSystemEventHandler()):	# Argumentos:
		self.observer = Observer() # Clase propia del watchdog para monitorear.		# 1. Directorio a monitorear.
		self.directory = directory # Instancia del directorio.						# 2. El handler a utilizar en caso de cambios en el directorio.
		self.handler = handler     # Instancia del handler.

	def execute(self):
		self.observer.schedule(self.handler, self.directory, recursive = True) # Programar monitoreo utilizando instancias anteriores.
		self.observer.start()												   # Recursive establece monitoreo de sub-folderes
		try:
			while True:
				time.sleep(1)
		except Exception:
			self.observer.stop()
		self.observer.join()

class myHandler(FileSystemEventHandler):
	def on_any_event(self, event):
		if event.event_type == "created":						# Verificar si el evento es de creación de archivo.
			for file in os.listdir(directory):					# Verificar en el directorio deseado.
				if fnmatch.fnmatch(file, "*.xls*"):				# Verificar si el archivo tiene extensión .xls*
					if fnmatch.fnmatch(file, "Master.xls*"):	# En caso sí, verificar si es el Master.xlsx (esto se hizo para que no mueva el archivo
						pass									# Master dado que se está ejecutando un for loop.
					else:
						# -------------- Copiando valores de celda de un workbook a otro -------------
						src_wb = xl.load_workbook(src_path+file)			# Cargar el workbook que se haya "creado".
						src_sheet_num = len(src_wb.sheetnames)				# Contar número de hojas.
						dest_wb = xl.load_workbook(src_path+"Master.xlsx")	# Cargar el Master.xlsx
						dest_sheet_num = len(dest_wb.sheetnames)			# Contar número de hojas del master.
						#print("sheet_num: "+ str(src_sheet_num))
						#print("dest_sheet_num: "+ str(dest_sheet_num))
						current_sheet = 0									# Iniciar variable current_sheet = 0.
						while current_sheet < src_sheet_num:				# Esta comparación se hace para verificar si ya se copiaron los datos de
																			# todas las hojas del archivo .xls* creado.
							src_ws = src_wb.worksheets[current_sheet]		# Seleccionar el worksheet del archivo creado, establecido por current_sheet
							dest_ws = dest_wb.worksheets[dest_sheet_num-1]	# Seleccionar el último worksheet vacío del Master.xlsx
							src_rows = src_ws.max_row						# Obtener el número máximo de filas del archivo creado.
							src_cols = src_ws.max_column					# Obtener el número máximo de columnas del archivo creado.
							for i in range(1, src_rows + 1):				# Ejecutar un for loop tanto para filas como columnas del archivo creado para
								for j in range(1, src_cols + 1):			# seleccionar todas las celdas utilizadas de la hoja y copiarlas.
									cells = src_ws.cell(row = i, column = j)
									dest_ws.cell(row = i, column = j).value = cells.value	# Pegar los datos de dichas celdas en el archivo Master.xlsx
							dest_sheet_num = dest_sheet_num + 1				# Incrementar el valor de dest_sheet_num para crear una hoja nueva en el Master.
							dest_wb.create_sheet("Sheet"+str(dest_sheet_num))	# Crear una hoja nueva en el Master para copiar los datos de la hoja siguiente.
							current_sheet = current_sheet + 1				# Incrementar current_sheet para determinar si ya se copiaron todas las hojas
						dest_wb.save(str(src_path+"Master.xlsx"))			# del archivo .xls* creado.
																			# Guardar el archivo Master.xlsx
						# ----------------------------------------------------------------------------
						print(file+" has been processed")
						shutil.move(src_path+file, xls_path+file)		# Mover el archivo creado que ya fue procesado al folder "Processed".
				elif not fnmatch.fnmatch(file, "*.xls*"):				# Condicional en caso que el archivo creado no sea tipo .xls*.
					if fnmatch.fnmatch(file, "*.py") or fnmatch.fnmatch(file, "Not_applicable") or fnmatch.fnmatch(file, "Processed"):
						pass											# Condicional para que el for ignore los folderes y el script de python durante
					else:												# la verificación y movida de archivos (si no se excluyen, genera un error).
						print(file+" is not applicable")				
						shutil.move(src_path+file, nonxls_path+file)	# Mover los archivo que no son de tipo .xls* al folder "Not_applicable"
		#if event =

if __name__ == "__main__":	# Requerimiento (el programa se ejecuta aquí), útil si se llega a importar a otro módulo.
	directory = input("Establish the directory to be watched: ")	# Le pide al usuario ingresar una dirección de directorio para monitorear.
	wt = watcher(directory, myHandler())							# Ejecuta la clase watcher usando el directorio obtenido y el handler que se programó.
	wt.execute()													# Inicia la observación del folder.